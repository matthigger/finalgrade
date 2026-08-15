"""End-to-end tests through the Config / CLI boundary.

These are deliberately written against the *public* boundary (a scope csv and
a config in, a grade dataframe out) rather than against Gradebook internals,
so that they survive refactoring of those internals.

Expected values are hand-computed and documented inline -- they are NOT
snapshots of current output.  Cases marked xfail(strict=True) are known bugs;
the marker is removed as each is fixed, and a strict xfail turns into a
failure the moment the behaviour starts working, so none can be forgotten.
"""
import os
import pathlib
import subprocess
import sys

import numpy as np
import pandas as pd
import pytest

from conftest import ASSIGN_STD, STUDENT_STD, write_scope
from gradescope_mean.__main__ import main, parser
from gradescope_mean.config import Config

CFG_BASE = """\
category:
  weight:
    hw: 1
    quiz: 1
"""


def known_bug(ref):
    """ marks behaviour that is specified here but not yet implemented

    strict=True means the marker itself fails the suite once the behaviour
    starts working, so no fix can silently leave a stale marker behind.
    """
    return pytest.mark.xfail(strict=True, reason=f'known bug: {ref}')


class TestPipelineGolden:
    """ the arithmetic, end to end, on a fixture whose values are known

              hw1  hw2  hw3  quiz1 | hw mean   quiz mean
    alice      10    8    6     10 | 24/30=.8       1.0
    bob        10   10   10      5 | 30/30=1.0       .5
    carol       0    6    6      6 | 12/30=.4        .6
    """

    def test_category_means(self, f_scope_std):
        _, df = Config(cat_weight_dict={'hw': 1, 'quiz': 1})(f_scope_std)

        np.testing.assert_allclose([.8, 1., .4], df['mean_hw'])
        np.testing.assert_allclose([1., .5, .6], df['mean_quiz'])
        # equal category weights -> plain average of the two category means
        np.testing.assert_allclose([.9, .75, .5], df['mean'])

    def test_letters(self, f_scope_std):
        _, df = Config(cat_weight_dict={'hw': 1, 'quiz': 1})(f_scope_std)
        # .90 -> A-,  .75 -> C (>= .73),  .50 -> E (below the .60 D- cut)
        assert df['letter'].tolist() == ['A-', 'C', 'E']

    def test_unequal_category_weights(self, f_scope_std):
        _, df = Config(cat_weight_dict={'hw': 3, 'quiz': 1})(f_scope_std)
        # alice (.8*3 + 1.0*1)/4 = .85
        # bob   (1.0*3 + .5*1)/4 = .875
        # carol (.4*3 + .6*1)/4  = .45
        np.testing.assert_allclose([.85, .875, .45], df['mean'])

    def test_drop_low(self, f_scope_std):
        _, df = Config(cat_weight_dict={'hw': 1, 'quiz': 1},
                       cat_drop_dict={'hw': 1})(f_scope_std)
        # alice drops the 6  -> 18/20 = .9
        # bob   drops a 10   -> 20/20 = 1.0
        # carol drops the 0  -> 12/20 = .6
        np.testing.assert_allclose([.9, 1., .6], df['mean_hw'])

    def test_late_penalty(self, f_scope_std):
        # bob is 1 day late, carol 3 days (2 on hw1 + 1 on hw2)
        # penalty = -.15 * unexcused / 3 assignments in category
        _, df = Config(
            cat_weight_dict={'hw': 1, 'quiz': 1},
            cat_late_dict={'hw': {'penalty_per_day': .15,
                                  'excuse_day': 0}})(f_scope_std)
        np.testing.assert_allclose([0, -1, -3], df['late days remain (hw)'])
        # alice .8, bob 1.0-.05=.95, carol .4-.15=.25
        np.testing.assert_allclose([.8, .95, .25], df['mean_hw'])

    def test_excuse_days(self, f_scope_std):
        _, df = Config(
            cat_weight_dict={'hw': 1, 'quiz': 1},
            cat_late_dict={'hw': {'penalty_per_day': .15,
                                  'excuse_day': 2}})(f_scope_std)
        # 2 free days each: alice +2 remaining, bob +1, carol -1
        np.testing.assert_allclose([2, 1, -1], df['late days remain (hw)'])
        # only carol is penalised, by 1 unexcused day
        np.testing.assert_allclose([.8, 1., .4 - .05], df['mean_hw'])

    def test_grade_full_columns(self, f_scope_std):
        _, df = Config(cat_weight_dict={'hw': 1, 'quiz': 1})(f_scope_std)
        for col in ('firstname', 'lastname', 'sid', 'mean', 'letter', 'hw1'):
            assert col in df.columns, col
        assert df.index.name == 'email'


class TestFeatureInteractions:
    """ the seams between features -- where the bugs actually live """

    def test_waive_also_waives_late_penalty(self, f_scope_std):
        """ waiving an assignment must waive its late days too (README) """
        _, df = Config(
            cat_weight_dict={'hw': 1, 'quiz': 1},
            waive_dict={'carol@u.edu': ['hw1']},
            cat_late_dict={'hw': {'penalty_per_day': .15,
                                  'excuse_day': 0}})(f_scope_std)
        # carol's hw1 is gone: mean_hw = 12/20 = .6, and only hw2's single
        # late day remains -> penalty .15*1/3 = .05  ->  .55
        assert df.loc['carol@u.edu', 'late days remain (hw)'] == -1
        assert df.loc['carol@u.edu', 'mean_hw'] == pytest.approx(.55)

    def test_waive_late_keeps_score_drops_penalty(self, f_scope_std):
        """ waive_late forgives lateness but keeps the score """
        _, df = Config(
            cat_weight_dict={'hw': 1, 'quiz': 1},
            late_waive_dict={'carol@u.edu': ['hw1']},
            cat_late_dict={'hw': {'penalty_per_day': .15,
                                  'excuse_day': 0}})(f_scope_std)
        # score still counts (mean_hw base stays .4) but hw1's 2 late days go
        assert df.loc['carol@u.edu', 'late days remain (hw)'] == -1
        assert df.loc['carol@u.edu', 'mean_hw'] == pytest.approx(.4 - .05)

    def test_email_list_prunes_before_penalty(self, f_scope_std):
        """ pruned students must not appear anywhere in the output """
        with pytest.warns(UserWarning):
            _, df = Config(
                cat_weight_dict={'hw': 1, 'quiz': 1},
                email_list=['alice@u.edu', 'bob@u.edu', 'ghost@u.edu'],
                cat_late_dict={'hw': {'penalty_per_day': .15,
                                      'excuse_day': 0}})(f_scope_std)
        assert df.index.tolist() == ['alice@u.edu', 'bob@u.edu']
        assert not df['mean'].isna().any()

    def test_substitute_then_exclude(self, tmp_path):
        """ substitute takes the max, exclude then removes the alternate """
        assignments = {'Quiz1': 10, 'Quiz1 v2': 10}
        students = [
            {'email': 'a@u.edu', 'scores': {'Quiz1': 4, 'Quiz1 v2': 9}},
            {'email': 'b@u.edu', 'scores': {'Quiz1': 8, 'Quiz1 v2': 3}},
        ]
        f = write_scope(tmp_path / 'scope.csv', assignments, students)
        # README's own substitute example uses names where one prefixes the
        # other, which trips the ambiguity warning from AssignmentList
        with pytest.warns(UserWarning, match='prefixes'):
            _, df = Config(sub_dict={'quiz1': ['quiz1v2']},
                           remove_list=['quiz1v2'])(f)
        # a takes the v2 score (.9), b keeps its own (.8)
        np.testing.assert_allclose([.9, .8], df['quiz1'])
        assert 'quiz1v2' not in df.columns
        np.testing.assert_allclose([.9, .8], df['mean'])

    def test_drop_low_respects_point_weights(self, tmp_path):
        """ drop-lowest is by percentage, mean is weighted by points """
        assignments = {'HW1': 10, 'HW2': 100}
        students = [{'email': 'a@u.edu', 'scores': {'HW1': 0, 'HW2': 50}}]
        f = write_scope(tmp_path / 'scope.csv', assignments, students)

        _, df = Config(cat_weight_dict={'hw': 1})(f)
        # (0*10 + .5*100) / 110
        assert df.loc['a@u.edu', 'mean_hw'] == pytest.approx(50 / 110)

        _, df = Config(cat_weight_dict={'hw': 1}, cat_drop_dict={'hw': 1})(f)
        # drops hw1 (0%), leaving the 100-pt assignment at 50%
        assert df.loc['a@u.edu', 'mean_hw'] == pytest.approx(.5)

    def test_complete_thresh_applied_after_substitute(self, tmp_path):
        """ substitution lifts completion above the threshold """
        assignments = {'Quiz1': 10, 'Quiz1 v2': 10}
        students = [
            {'email': 'a@u.edu', 'scores': {'Quiz1 v2': 9}},
            {'email': 'b@u.edu', 'scores': {'Quiz1 v2': 8}},
        ]
        f = write_scope(tmp_path / 'scope.csv', assignments, students)
        # nobody submitted quiz1 itself; after substitution everyone has a
        # score, so a 60% completion threshold must not drop it
        with pytest.warns(UserWarning, match='prefixes'):
            _, df = Config(sub_dict={'quiz1': ['quiz1v2']},
                           remove_list=['quiz1v2'],
                           exclude_complete_thresh=.6)(f)
        assert 'quiz1' in df.columns


class TestAdversarialInput:
    """ malformed / unusual gradescope exports """

    def test_zero_point_assignment(self, tmp_path):
        """ a 0-point assignment must not poison the mean with inf/nan """
        assignments = {'HW1': 10, 'Survey': 0}
        students = [
            {'email': 'a@u.edu', 'scores': {'HW1': 8, 'Survey': 1}},
            {'email': 'b@u.edu', 'scores': {'HW1': 9, 'Survey': 0}},
        ]
        f = write_scope(tmp_path / 'scope.csv', assignments, students)
        with pytest.warns(UserWarning, match='0 points'):
            _, df = Config()(f)
        np.testing.assert_allclose([.8, .9], df['mean'])
        assert not df['mean'].isna().any()

    def test_blank_lateness_cell(self, tmp_path):
        """ an empty lateness cell means 'not late', not a crash """
        f = tmp_path / 'scope.csv'
        f.write_text(
            'First Name,Last Name,SID,Email,Sections,'
            'HW1,HW1 - Max Points,HW1 - Submission Time,'
            'HW1 - Lateness (H:M:S)\n'
            'a,b,1S,a@u.edu,sec,8,10,,\n')
        gb, df = Config()(f)
        assert gb.df_lateday.loc['a@u.edu', 'hw1'] == 0
        assert df.loc['a@u.edu', 'mean'] == pytest.approx(.8)

    def test_duplicate_email_raises(self, tmp_path):
        """ two rows for one email cannot be silently merged """
        students = list(STUDENT_STD) + [dict(STUDENT_STD[0], scores={
            'HW1': 0, 'HW2': 0, 'HW3': 0, 'Quiz1': 0})]
        f = write_scope(tmp_path / 'scope.csv', ASSIGN_STD, students)
        with pytest.raises(ValueError, match='(?i)duplicate'):
            Config()(f)

    def test_extra_metadata_column(self, tmp_path):
        """ exports carry varying numbers of metadata columns """
        f = write_scope(
            tmp_path / 'scope.csv', ASSIGN_STD, STUDENT_STD,
            meta_header=['First Name', 'Last Name', 'SID', 'Email',
                         'Sections', 'CRN'])
        _, df = Config(cat_weight_dict={'hw': 1, 'quiz': 1})(f)
        # the extra column must be metadata, never mistaken for an assignment
        np.testing.assert_allclose([.9, .75, .5], df['mean'])

    def test_fewer_metadata_columns(self, tmp_path):
        """ an export with no SID column still works """
        f = write_scope(
            tmp_path / 'scope.csv', ASSIGN_STD, STUDENT_STD,
            meta_header=['First Name', 'Last Name', 'Email', 'Sections'])
        _, df = Config(cat_weight_dict={'hw': 1, 'quiz': 1})(f)
        np.testing.assert_allclose([.9, .75, .5], df['mean'])

    def test_score_above_max_points(self, tmp_path):
        """ extra credit above max points is preserved, not clipped """
        assignments = {'HW1': 10}
        students = [{'email': 'a@u.edu', 'scores': {'HW1': 12}}]
        f = write_scope(tmp_path / 'scope.csv', assignments, students)
        _, df = Config()(f)
        assert df.loc['a@u.edu', 'mean'] == pytest.approx(1.2)


class TestConfigValidationEndToEnd:
    """ a bad config must fail loudly, never produce plausible wrong grades """

    def test_grade_thresh_on_0_100_scale_raises(self, f_scope_std):
        """ '93: A' instead of '.93: A' must not silently fail everyone """
        with pytest.raises(ValueError, match='(?i)grade_thresh'):
            Config(cat_weight_dict={'hw': 1},
                   grade_thresh={93: 'A', 80: 'B', 0: 'E'})

    def test_late_penalty_on_unknown_category_raises(self, f_scope_std):
        """ a typo'd late_penalty category must not be silently ignored """
        with pytest.raises(ValueError, match='(?i)late_penalty'):
            Config(cat_weight_dict={'hw': 1, 'quiz': 1},
                   cat_late_dict={'homework': {'penalty_per_day': .15}})

    def test_all_zero_weights_raises(self):
        with pytest.raises(ValueError, match='(?i)weight'):
            Config(cat_weight_dict={'hw': 0, 'quiz': 0})

    def test_drop_low_on_unknown_category_raises(self):
        with pytest.raises(ValueError, match='(?i)drop_low'):
            Config(cat_weight_dict={'hw': 1}, cat_drop_dict={'quiz': 1})

    def test_category_matching_no_assignment_raises(self, f_scope_std):
        """ a weight category that matches nothing is a config error """
        with pytest.raises(ValueError, match='(?i)(match|category)'):
            Config(cat_weight_dict={'hw': 1, 'lab': 1})(f_scope_std)

    def test_penalty_per_day_on_0_100_scale_raises(self):
        """ '15' instead of '.15' is a 100x penalty: enough to zero a class """
        with pytest.raises(ValueError, match='(?i)penalty_per_day'):
            Config(cat_weight_dict={'hw': 1},
                   cat_late_dict={'hw': {'penalty_per_day': 15}})

    def test_negative_penalty_per_day_raises(self):
        """ a negative penalty would award credit for being late """
        with pytest.raises(ValueError, match='(?i)penalty_per_day'):
            Config(cat_weight_dict={'hw': 1},
                   cat_late_dict={'hw': {'penalty_per_day': -.15}})

    def test_late_penalty_without_penalty_per_day_raises(self):
        with pytest.raises(ValueError, match='(?i)penalty_per_day'):
            Config(cat_weight_dict={'hw': 1},
                   cat_late_dict={'hw': {'excuse_day': 3}})

    def test_negative_excuse_day_raises(self):
        with pytest.raises(ValueError, match='(?i)excuse_day'):
            Config(cat_weight_dict={'hw': 1},
                   cat_late_dict={'hw': {'penalty_per_day': .15,
                                         'excuse_day': -3}})

    def test_excuse_day_offset_must_be_numbers(self):
        with pytest.raises(ValueError, match='(?i)excuse_day_offset'):
            Config(cat_weight_dict={'hw': 1},
                   cat_late_dict={'hw': {
                       'penalty_per_day': .15,
                       'excuse_day_offset': {'a@u.edu': 'three'}}})

    def test_late_penalty_block_must_be_a_mapping(self):
        """ 'late_penalty: {hw: .15}' is settings-shaped, but isn't """
        with pytest.raises(ValueError, match='(?i)late_penalty'):
            Config(cat_weight_dict={'hw': 1}, cat_late_dict={'hw': .15})


class TestConfigKeyChecking:
    """ a key we don't recognize is a typo, and a typo we ignore is a
    grading policy silently not applied """

    def test_unknown_top_level_key_raises(self, write_config):
        f = write_config('catagory:\n  weight:\n    hw: 1\n')
        with pytest.raises(ValueError, match='catagory'):
            Config.from_file(f)

    def test_unknown_key_suggests_the_real_one(self, write_config):
        f = write_config('catagory:\n  weight:\n    hw: 1\n')
        with pytest.raises(ValueError, match='did you mean category'):
            Config.from_file(f)

    def test_unknown_nested_key_raises(self, write_config):
        """ 'weights' under category was silently ignored """
        f = write_config('category:\n  weights:\n    hw: 1\n')
        with pytest.raises(ValueError, match='category/weights'):
            Config.from_file(f)

    def test_unknown_assignments_key_raises(self, write_config):
        f = write_config('assignments:\n  exclude_complete_threshold: .6\n')
        with pytest.raises(ValueError,
                           match='exclude_complete_thresh'):
            Config.from_file(f)

    def test_section_given_a_scalar_raises(self, write_config):
        """ 'category: 50' can't be read, and must not be skipped """
        f = write_config('category: 50\n')
        with pytest.raises(ValueError, match='(?i)category.*mapping'):
            Config.from_file(f)

    def test_unknown_late_penalty_setting_raises(self, write_config):
        """ this block is passed to get_late_penalty() as kwargs: an unknown
        key used to surface as a TypeError, escaping the CLI's handler """
        f = write_config("""\
category:
  weight:
    hw: 1
  late_penalty:
    hw:
      penalty_per_day: .15
      excuse_days: 3
""")
        with pytest.raises(ValueError, match='excuse_day'):
            Config.from_file(f)

    def test_user_names_below_a_leaf_are_not_checked(self, write_config,
                                                     f_scope_std):
        """ categories, assignments and emails are the user's own words """
        f = write_config("""\
category:
  weight:
    hw: 1
    quiz: 1
  drop_low:
    hw: 1
  late_penalty:
    hw:
      penalty_per_day: .15
      excuse_day: 1
      excuse_day_offset:
        alice@u.edu: 2
assignments:
  exclude_complete_thresh: .1
  substitute:
    hw1:
      - hw2
waive:
  alice@u.edu: hw3
waive_late:
  bob@u.edu: hw1
grade_thresh:
  .9: A
  0: E
email_list:
  - alice@u.edu
  - bob@u.edu
  - carol@u.edu
""")
        # every section populated at once: loads and runs
        _, df = Config.from_file(f)(f_scope_std)
        assert len(df) == 3

    def test_default_shipped_config_passes_key_check(self):
        """ the config every new user starts from must not trip its own check
        """
        from gradescope_mean.config import F_CONFIG_DEFAULT
        Config.from_file(F_CONFIG_DEFAULT)


class TestScalarInsteadOfList:
    """ a bare string is a list of one, not a list of characters """

    def test_exclude_as_scalar(self, write_config, f_scope_std):
        """ 'exclude: quiz1' (a forgotten '-') removed hw1 as well, because
        'quiz1' iterated to ['q', 'u', 'i', 'z', '1'] """
        f = write_config('assignments:\n  exclude: quiz1\n')
        _, df = Config.from_file(f)(f_scope_std)
        # hw1, hw2, hw3 remain, 10 pts each: 24/30, 30/30, 12/30
        np.testing.assert_allclose([.8, 1., .4], df['mean'])

    def test_exclude_as_comma_separated_string(self, write_config,
                                               f_scope_std):
        f = write_config('assignments:\n  exclude: quiz1, hw3\n')
        _, df = Config.from_file(f)(f_scope_std)
        # hw1, hw2 remain: 18/20, 20/20, 6/20
        np.testing.assert_allclose([.9, 1., .3], df['mean'])

    def test_exclude_as_list_is_unchanged(self, write_config, f_scope_std):
        f = write_config('assignments:\n  exclude:\n    - quiz1\n')
        _, df = Config.from_file(f)(f_scope_std)
        np.testing.assert_allclose([.8, 1., .4], df['mean'])

    def test_email_list_as_scalar(self, write_config, f_scope_std):
        f = write_config('email_list: alice@u.edu\n')
        _, df = Config.from_file(f)(f_scope_std)
        assert df.index.tolist() == ['alice@u.edu']

    def test_exclude_of_a_mapping_raises(self, write_config):
        """ a list of mappings is a shape we can't read: say so """
        f = write_config('assignments:\n  exclude:\n    - quiz1: yes\n')
        with pytest.raises(ValueError, match='(?i)exclude'):
            Config.from_file(f)


class TestDeterminism:
    def test_output_row_order_is_stable(self, tmp_path):
        """ same input -> byte-identical output, whatever the hash seed """
        f_scope = write_scope(tmp_path / 'scope.csv', ASSIGN_STD, STUDENT_STD)
        script = (
            'import sys, warnings; warnings.simplefilter("ignore");'
            f'sys.path.insert(0, {str(pathlib.Path.cwd())!r});'
            'from gradescope_mean.config import Config;'
            f'_, df = Config(email_list={[s["email"] for s in STUDENT_STD]!r})'
            f'({str(f_scope)!r});'
            'print("ORDER:" + ",".join(df.index))')
        out_set = set()
        for seed in ('0', '1', '2', '3', '5'):
            env = dict(os.environ, PYTHONHASHSEED=seed)
            res = subprocess.run([sys.executable, '-c', script],
                                 capture_output=True, text=True, env=env)
            assert res.returncode == 0, res.stderr
            line, = [l for l in res.stdout.splitlines()
                     if l.startswith('ORDER:')]
            out_set.add(line)
        assert len(out_set) == 1, f'row order varies by hash seed: {out_set}'

    def test_row_order_follows_input(self, f_scope_std):
        """ output preserves gradescope's row order """
        _, df = Config(email_list=['carol@u.edu', 'alice@u.edu'])(f_scope_std)
        assert df.index.tolist() == ['alice@u.edu', 'carol@u.edu']


class TestCLI:
    def _write_cfg(self, tmp_path, text=CFG_BASE):
        f = tmp_path / 'config.yaml'
        f.write_text(text)
        return str(f)

    def test_grade_writes_expected_values(self, tmp_path, f_scope_std):
        f_cfg = self._write_cfg(tmp_path)
        main(parser.parse_args(
            ['grade', str(f_scope_std), '--config', f_cfg, '-q']))
        df = pd.read_csv(tmp_path / 'grade_full.csv', index_col='email')
        np.testing.assert_allclose([.9, .75, .5], df['mean'])
        assert df['letter'].tolist() == ['A-', 'C', 'E']

    def test_quiet_suppresses_stdout(self, tmp_path, f_scope_std, capsys):
        f_cfg = self._write_cfg(tmp_path)
        main(parser.parse_args(
            ['grade', str(f_scope_std), '--config', f_cfg, '-q']))
        assert capsys.readouterr().out == ''

    def test_late_csv_uses_configured_grace(self, tmp_path, f_scope_std):
        """ the exported late days must match the ones actually penalised """
        f_cfg = self._write_cfg(tmp_path, CFG_BASE + """\
  late_penalty:
    hw:
      penalty_per_day: .15
      excuse_day: 0
      grace_period_minutes: 1500
""")
        main(parser.parse_args(
            ['grade', str(f_scope_std), '--config', f_cfg,
             '--late_csv', 'late.csv', '-q']))
        df_late = pd.read_csv(tmp_path / 'late.csv', index_col=0)
        # 1500 min = 25h of grace, so bob's 24h hw1 is not late at all
        assert df_late.loc['bob@u.edu', 'hw1'] == 0

    def test_per_student_files_are_unique(self, tmp_path):
        """ students sharing a name must not overwrite each other's file """
        students = [
            {'email': 'a@u.edu', 'first': 'sam', 'last': 'smith',
             'scores': {'HW1': 10}},
            {'email': 'b@u.edu', 'first': 'sam', 'last': 'smith',
             'scores': {'HW1': 5}},
        ]
        f = write_scope(tmp_path / 'scope.csv', {'HW1': 10}, students)
        f_cfg = self._write_cfg(tmp_path, 'category:\n  weight: null\n')
        main(parser.parse_args(
            ['grade', str(f), '--config', f_cfg, '--per_student', '-q']))
        assert len(list((tmp_path / 'per_student').glob('*.csv'))) == 2


class TestExporters:
    def test_canvas_cli_end_to_end(self, tmp_path):
        """ the documented canvas command, on an export whose section column
        is named 'section_name' rather than 'Sections' """
        f_scope = write_scope(
            tmp_path / 'scope.csv', ASSIGN_STD, STUDENT_STD,
            meta_header=['First Name', 'Last Name', 'SID', 'Email',
                         'section_name'])
        f_cfg = tmp_path / 'config.yaml'
        f_cfg.write_text(CFG_BASE)
        main(parser.parse_args(
            ['grade', str(f_scope), '--config', str(f_cfg), '-q']))

        f_canvas = tmp_path / 'canvas.csv'
        pd.DataFrame({
            'Student': ['Anders, Alice', 'Baker, Bob'],
            'ID': [100, 101],
            'SIS User ID': ['001S', '002S'],
            'SIS Login ID': ['alice@u.edu', 'bob@u.edu'],
            'Section': ['sec01', 'sec01'],
            'Placeholder': [0, 0]}).to_csv(f_canvas, index=False)

        main(parser.parse_args(
            ['canvas', str(tmp_path / 'grade_full.csv'), str(f_canvas), '-q']))
        out_list = [p for p in tmp_path.glob('canvas*.csv')
                    if p.name != 'canvas.csv']
        assert len(out_list) == 1
        df = pd.read_csv(out_list[0])
        assert 'mean' in df.columns
        np.testing.assert_allclose([.9, .75], df['mean'])

    def test_canvas_merge_does_not_mutate_arguments(self, tmp_path):
        from gradescope_mean.canvas.canvas import canvas_merge
        f_scope = write_scope(tmp_path / 'scope.csv', ASSIGN_STD, STUDENT_STD)
        _, df_grade = Config(cat_weight_dict={'hw': 1, 'quiz': 1})(f_scope)
        df_grade = df_grade.reset_index()

        f_canvas = tmp_path / 'canvas.csv'
        pd.DataFrame({'Student': ['a'], 'ID': [1], 'SIS User ID': ['001S'],
                      'SIS Login ID': ['alice@u.edu'], 'Section': ['s'],
                      'Placeholder': [0]}).to_csv(f_canvas, index=False)

        cols_before = list(df_grade.columns)
        del_list = ['letter']
        canvas_merge(f_canvas=str(f_canvas), df_grade=df_grade,
                     del_col_list=del_list, scale100=False)
        assert list(df_grade.columns) == cols_before
        assert del_list == ['letter']
        # and it must be safe to call twice
        canvas_merge(f_canvas=str(f_canvas), df_grade=df_grade,
                     scale100=False)

    def test_banner_cli_without_crn(self, tmp_path, f_scope_std):
        f_cfg = tmp_path / 'config.yaml'
        f_cfg.write_text(CFG_BASE)
        main(parser.parse_args(
            ['grade', str(f_scope_std), '--config', str(f_cfg), '-q']))
        main(parser.parse_args(
            ['banner', str(tmp_path / 'grade_full.csv'), '202410', '-q']))
        assert len(list(tmp_path.glob('*banner*.xlsx'))) == 1

    def test_banner_preserves_leading_zeros(self, tmp_path):
        import openpyxl
        students = [{'email': 'a@u.edu', 'sid': '0001234S',
                     'scores': {'HW1': 10}}]
        f_scope = write_scope(tmp_path / 'scope.csv', {'HW1': 10}, students)
        f_cfg = tmp_path / 'config.yaml'
        f_cfg.write_text('category:\n  weight: null\n')
        main(parser.parse_args(
            ['grade', str(f_scope), '--config', str(f_cfg), '-q']))
        main(parser.parse_args(
            ['banner', str(tmp_path / 'grade_full.csv'), '202410', '-q']))

        f_xlsx = next(tmp_path.glob('*banner*.xlsx'))
        ws = openpyxl.load_workbook(f_xlsx).active
        header = [c.value for c in ws[1]]
        col = header.index('Student ID')
        # read the raw cell: pd.read_excel would coerce this to an int and
        # hide exactly the bug we are checking for
        assert ws[2][col].value == '000001234'
