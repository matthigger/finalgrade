import copy
import pathlib

import openpyxl
import pandas as pd
import pytest
from conftest import ASSIGN_STD, STUDENT_STD, write_scope

import finalgrade
from finalgrade.__main__ import main, parser
from finalgrade.errors import FinalgradeError
from finalgrade.policy import Policy

test_folder = pathlib.Path(finalgrade.__file__).parents[1] / 'test'

# the section names gradescope actually writes: a course, a CRN, a title and
# a term, run together.  nobody is going to type one of these out
SEC_TUP = ('cs2810-34240-mathematics-of-data-models-sec-01-spring-2022',
           'cs2810-34241-mathematics-of-data-models-sec-02-spring-2022',
           'cs2810-34242-mathematics-of-data-models-sec-03-spring-2022')


@pytest.fixture
def grade_full_csv(tmp_path):
    """Create a grade_full.csv from test data"""
    policy = Policy()
    _, df_grade_full = policy(f_scope=test_folder / 'scope.csv')
    f = tmp_path / 'grade_full.csv'
    df_grade_full.to_csv(f)
    return str(f)


@pytest.fixture
def grade_full_multi(tmp_path):
    """ a grade_full.csv whose three students sit in three sections """
    student_list = copy.deepcopy(STUDENT_STD)
    for stud, sec in zip(student_list, SEC_TUP):
        stud['section'] = sec

    f_scope = write_scope(tmp_path / 'scope.csv', ASSIGN_STD, student_list)
    _, df_grade_full = Policy()(f_scope=str(f_scope))
    f = tmp_path / 'grade_full.csv'
    df_grade_full.to_csv(f)
    return str(f)


def _banner_args(grade_full_csv, *extra):
    return parser.parse_args(['banner', grade_full_csv, '202310', *extra])


def _read_out(tmp_path, **kwarg_dict):
    """ the one workbook the banner subcommand just wrote """
    xlsx_list = list(tmp_path.glob('*banner*.xlsx'))
    assert len(xlsx_list) == 1
    return pd.read_excel(xlsx_list[0], **kwarg_dict)


def _cli_error(capsys, *arg_tup):
    """ what the banner subcommand printed before giving up

    the cli turns a FinalgradeError into one line and exit(2), which is the
    thing a user sees -- so that is what these assert against.
    """
    with pytest.raises(SystemExit) as exc_info:
        main(_banner_args(*arg_tup))
    assert exc_info.value.code == 2
    return capsys.readouterr().err


class TestBanner:
    def test_basic_banner_export(self, grade_full_csv, tmp_path):
        main(_banner_args(grade_full_csv, '-c', '12345', '-q'))

        df = _read_out(tmp_path)
        assert 'Term Code' in df.columns
        assert 'CRN0' in df.columns
        assert 'Student ID' in df.columns
        assert 'sid' not in df.columns

    def test_multiple_crns(self, grade_full_csv, tmp_path):
        main(_banner_args(grade_full_csv, '-c', '11111', '-c', '22222', '-q'))

        df = _read_out(tmp_path)
        assert 'CRN0' in df.columns
        assert 'CRN1' in df.columns

    def test_no_crn(self, grade_full_csv, tmp_path):
        """ -c is optional """
        main(_banner_args(grade_full_csv, '-q'))
        assert len(list(tmp_path.glob('*banner*.xlsx'))) == 1

    def test_student_id_keeps_leading_zeros(self, grade_full_csv, tmp_path):
        """ banner ids are 9 digit strings; a leading zero is significant

        read the raw cells: pd.read_excel coerces these to ints, which would
        hide exactly the bug being checked for.
        """
        main(_banner_args(grade_full_csv, '-q'))

        f_xlsx = next(tmp_path.glob('*banner*.xlsx'))
        ws = openpyxl.load_workbook(f_xlsx).active
        header_list = [c.value for c in ws[1]]
        idx = header_list.index('Student ID')

        for row in list(ws.iter_rows(min_row=2)):
            value = row[idx].value
            assert isinstance(value, str)
            assert value.isdigit()
            assert len(value) >= 9

        # scope.csv's first student is 0123456789S: the trailing S goes, the
        # leading zero stays, and zfill only pads ids shorter than 9
        assert ws[2][idx].value == '0123456789'


class TestLetterOnly:
    """ banner reads a final grade, so that is what it is handed """

    def test_only_the_columns_banner_reads(self, grade_full_csv, tmp_path):
        main(_banner_args(grade_full_csv, '-c', '12345', '-q'))

        df = _read_out(tmp_path)
        assert list(df.columns) == ['Term Code', 'CRN0', 'Student ID',
                                    'Final Grade']

    def test_the_final_grade_is_the_letter(self, grade_full_csv, tmp_path):
        main(_banner_args(grade_full_csv, '-q'))

        df_out = _read_out(tmp_path)
        df_grade = pd.read_csv(grade_full_csv)
        assert list(df_out['Final Grade']) == list(df_grade['letter'])

    def test_full_keeps_every_grade_column(self, grade_full_csv, tmp_path):
        main(_banner_args(grade_full_csv, '-c', '12345', '--full', '-q'))

        df = _read_out(tmp_path)
        assert 'mean' in df.columns
        assert 'letter' in df.columns
        assert 'hw1' in df.columns
        # banner's own fields still lead, in the order the trimmed workbook
        # uses: the toggle changes how much comes along, not the shape of it
        assert list(df.columns)[:3] == ['Term Code', 'CRN0', 'Student ID']

    def test_grades_with_no_letter_are_refused(self, grade_full_csv,
                                               tmp_path):
        from finalgrade.banner import to_banner

        df = pd.read_csv(grade_full_csv, dtype={'sid': str})
        del df['letter']

        with pytest.raises(FinalgradeError, match='no .letter. column'):
            to_banner(df, term_code='202310')


class TestSectionCrn:
    def test_one_crn_column_filled_per_section(self, grade_full_multi,
                                               tmp_path):
        main(_banner_args(grade_full_multi,
                          '-s', f'{SEC_TUP[0]}=11111',
                          '-s', f'{SEC_TUP[1]}=22222',
                          '-s', f'{SEC_TUP[2]}=33333', '-q'))

        df = _read_out(tmp_path, dtype=str)
        # exactly 'CRN', which banner recognises and pre-selects
        assert 'CRN' in df.columns
        assert 'CRN0' not in df.columns
        assert sorted(df['CRN']) == ['11111', '22222', '33333']

    def test_the_crn_lands_on_the_right_student(self, grade_full_multi,
                                                tmp_path):
        with pytest.warns(UserWarning, match='no CRN'):
            main(_banner_args(grade_full_multi, '-s', 'sec-02=22222', '-q'))

        df = _read_out(tmp_path, dtype=str)
        # bob is the second student, and the only one in sec-02
        assert list(df['Student ID']) == ['000000002']
        assert list(df['CRN']) == ['22222']

    def test_a_section_is_matched_by_the_part_you_typed(self, grade_full_multi,
                                                        tmp_path):
        """ nobody types the whole gradescope section name """
        main(_banner_args(grade_full_multi, '-s', 'sec-01=11111',
                          '-s', '34241=22222', '-s', 'sec-03=33333', '-q'))

        df = _read_out(tmp_path, dtype=str)
        assert sorted(df['CRN']) == ['11111', '22222', '33333']

    def test_an_unmapped_section_is_left_out_loudly(self, grade_full_multi,
                                                    tmp_path):
        with pytest.warns(UserWarning, match='sec-02'):
            main(_banner_args(grade_full_multi, '-s', 'sec-01=11111',
                              '-s', 'sec-03=33333', '-q'))

        df = _read_out(tmp_path, dtype=str)
        assert sorted(df['CRN']) == ['11111', '33333']

    def test_a_blank_crn_leaves_its_section_out(self, grade_full_multi,
                                                tmp_path):
        """ the web form sends a box per section, and most stay empty """
        from finalgrade.banner import to_banner

        df = pd.read_csv(grade_full_multi, dtype={'sid': str})
        with pytest.warns(UserWarning, match='2 students'):
            df_out = to_banner(df, term_code='202310',
                               crn_dict={SEC_TUP[0]: '11111',
                                         SEC_TUP[1]: '', SEC_TUP[2]: None})

        assert list(df_out['CRN']) == ['11111']

    def test_a_section_nobody_is_in_is_an_error(self, grade_full_multi,
                                                capsys):
        text = _cli_error(capsys, grade_full_multi, '-s', 'sec-09=99999',
                          '-q')
        assert 'no section matches' in text

    def test_a_near_miss_names_the_real_section(self, grade_full_multi,
                                                capsys):
        text = _cli_error(capsys, grade_full_multi,
                          '-s', f'{SEC_TUP[0]}x=11111', '-q')
        assert 'did you mean' in text
        assert SEC_TUP[0] in text

    def test_a_fragment_matching_two_sections_is_an_error(self,
                                                          grade_full_multi,
                                                          capsys):
        text = _cli_error(capsys, grade_full_multi, '-s', 'cs2810=11111',
                          '-q')
        assert 'matches 3 sections' in text

    def test_two_names_for_one_section_disagreeing_is_an_error(
            self, grade_full_multi):
        from finalgrade.banner import to_banner

        df = pd.read_csv(grade_full_multi, dtype={'sid': str})
        with pytest.raises(FinalgradeError, match='two CRNs'):
            to_banner(df, term_code='202310',
                      crn_dict={'sec-01': '11111', '34240': '22222'})

    def test_one_section_twice_on_the_command_line_is_an_error(
            self, grade_full_multi, capsys):
        text = _cli_error(capsys, grade_full_multi, '-s', 'sec-01=11111',
                          '-s', 'sec-01=22222', '-q')
        assert 'only have one CRN' in text

    def test_no_section_mapped_at_all_is_an_error(self, grade_full_multi):
        from finalgrade.banner import to_banner

        df = pd.read_csv(grade_full_multi, dtype={'sid': str})
        with pytest.raises(FinalgradeError, match='no section was given'):
            to_banner(df, term_code='202310', crn_dict={'sec-01': ''})

    def test_both_kinds_of_crn_is_an_error(self, grade_full_multi, capsys):
        text = _cli_error(capsys, grade_full_multi, '-c', '11111',
                          '-s', 'sec-01=22222', '-q')
        assert 'use one or the other' in text

    def test_a_gradebook_with_no_sections_says_so(self, grade_full_csv,
                                                  tmp_path):
        from finalgrade.banner import to_banner

        df = pd.read_csv(grade_full_csv, dtype={'sid': str})
        del df['section_name']

        with pytest.raises(FinalgradeError, match='name no section'):
            to_banner(df, term_code='202310', crn_dict={'sec-01': '11111'})

    def test_section_crn_wants_an_equals_sign(self, grade_full_multi):
        with pytest.raises(SystemExit):
            _banner_args(grade_full_multi, '-s', 'sec-01')

    def test_the_mapping_stays_out_of_the_policy(self, grade_full_multi,
                                                 tmp_path):
        """ where grades go is not how they were earned """
        main(_banner_args(grade_full_multi, '-s', 'sec-01=11111',
                          '-s', 'sec-02=22222', '-s', 'sec-03=33333', '-q'))

        for f in tmp_path.glob('*.yaml'):
            assert '11111' not in f.read_text()


class TestSectionList:
    def test_it_finds_the_sections(self, grade_full_multi):
        from finalgrade.banner import section_list

        df = pd.read_csv(grade_full_multi)
        assert section_list(df) == sorted(SEC_TUP)

    def test_no_section_column_is_no_sections(self):
        from finalgrade.banner import section_list

        assert section_list(pd.DataFrame({'sid': ['1']})) == []

    def test_the_canvas_spelling_is_read_too(self):
        from finalgrade.banner import section_list

        assert section_list(pd.DataFrame({'sections': ['a', 'b', 'a']})) == \
            ['a', 'b']

    def test_blank_sections_are_not_sections(self):
        from finalgrade.banner import section_list

        assert section_list(pd.DataFrame({'sections': ['a', '', None]})) == \
            ['a']


class TestBannerId:
    def test_the_s_suffix_goes(self):
        from finalgrade.banner import banner_id

        assert banner_id('1234567S') == '001234567'

    def test_a_short_id_is_padded(self):
        from finalgrade.banner import banner_id

        assert banner_id('42') == '000000042'

    def test_an_int_id_survives(self):
        """ read_scope leaves sid uncast, so it is sometimes a number """
        from finalgrade.banner import banner_id

        assert banner_id(1234567) == '001234567'


class TestBannerWeb:
    """ the browser build writes the same workbook, from the same function """

    @pytest.fixture
    def csv_text(self, f_scope_std):
        return f_scope_std.read_text()

    @pytest.fixture
    def csv_text_multi(self, tmp_path):
        student_list = copy.deepcopy(STUDENT_STD)
        for stud, sec in zip(student_list, SEC_TUP):
            stud['section'] = sec
        return write_scope(tmp_path / 'multi.csv', ASSIGN_STD,
                           student_list).read_text()

    def test_builds_a_workbook(self, csv_text):
        import base64
        import io

        from finalgrade import web

        res = web.banner_export(csv_text, '', '202310', '["12345"]')

        assert res['ok'], res.get('error')
        data = base64.b64decode(res['xlsx_b64'])
        # a zip container, which is what an xlsx is
        assert data[:2] == b'PK'

        # read as text: pandas coerces these to ints otherwise, which would
        # hide whether they were written as digits at all
        df = pd.read_excel(io.BytesIO(data), dtype=str)
        assert list(df['Term Code']) == ['202310'] * 3
        assert list(df['CRN0']) == ['12345'] * 3
        assert 'sid' not in df.columns

    def test_it_matches_the_command_line(self, csv_text, f_scope_std,
                                         tmp_path):
        import base64
        import io

        from finalgrade import web
        from finalgrade.banner import to_banner

        res = web.banner_export(csv_text, '', '202310', '["12345"]')
        df_web = pd.read_excel(io.BytesIO(base64.b64decode(res['xlsx_b64'])),
                               dtype={'Student ID': str})

        _, df_grade = Policy()(str(f_scope_std))
        f = tmp_path / 'grade_full.csv'
        df_grade.to_csv(f)
        df_cli = to_banner(pd.read_csv(f, dtype={'sid': str}),
                           term_code='202310', crn_list=['12345'])

        assert list(df_web['Student ID']) == list(df_cli['Student ID'])
        assert list(df_web['Final Grade']) == list(df_cli['Final Grade'])

    def test_a_mapping_is_read_as_a_mapping(self, csv_text_multi):
        import base64
        import io

        from finalgrade import web

        res = web.banner_export(
            csv_text_multi, '', '202310',
            f'{{"{SEC_TUP[0]}": "11111", "{SEC_TUP[1]}": "22222", '
            f'"{SEC_TUP[2]}": "33333"}}')

        assert res['ok'], res.get('error')
        df = pd.read_excel(io.BytesIO(base64.b64decode(res['xlsx_b64'])),
                           dtype=str)
        assert sorted(df['CRN']) == ['11111', '22222', '33333']
        assert res['n_row'] == 3

    def test_a_section_left_blank_is_a_warning_not_an_error(self,
                                                            csv_text_multi):
        from finalgrade import web

        res = web.banner_export(
            csv_text_multi, '', '202310', f'{{"{SEC_TUP[0]}": "11111"}}')

        assert res['ok'], res.get('error')
        assert res['n_row'] == 1
        assert any('no CRN' in w for w in res['warn_list'])

    def test_no_crn_at_all_is_a_message(self, csv_text_multi):
        from finalgrade import web

        res = web.banner_export(csv_text_multi, '', '202310', '{}')

        assert not res['ok']
        assert 'no section was given a CRN' in res['error']

    def test_no_term_code_is_refused(self, csv_text):
        from finalgrade import web

        res = web.banner_export(csv_text, '', '   ', '[]')

        assert not res['ok']
        assert 'term code' in res['error']

    def test_a_broken_config_is_a_message(self, csv_text):
        from finalgrade import web

        res = web.banner_export(
            csv_text, 'category:\n  weight:\n    nope: 1\n', '202310', '[]')

        assert not res['ok']
        assert 'nope' in res['error']

    def test_crns_are_optional(self, csv_text):
        from finalgrade import web

        assert web.banner_export(csv_text, '', '202310', '[]')['ok']

    def test_load_csv_names_the_sections(self, csv_text_multi):
        from finalgrade import web

        info = web.load_csv(csv_text_multi, 'multi.csv')

        assert info['ok'], info.get('error')
        assert info['section_list'] == sorted(SEC_TUP)
